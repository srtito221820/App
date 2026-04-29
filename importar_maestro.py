"""
Importa datos del Maestro_Telas de los Excel de control a la tabla maestro_telas.

Uso:
    python importar_maestro.py

Detecta automaticamente los dos archivos conocidos (Ritex y Neocxela) y los
vincula con sus proveedores por nombre. Las telas MORLEY y RIBB/REEB se marcan
con cuenta_piezas=False para que el control de piezas las ignore.
"""
import sys
import io
import os

# Reconfigurar stdout para evitar errores cp1252 en Windows
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
except Exception:
    pass

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from app import app
from models import db, MaestroTela, Proveedor


# Archivo -> nombre (o fragmento) del proveedor asociado
FUENTES = [
    (r'C:\Users\MGaravaglia\Desktop\Control_Telas Ritex.xlsx',   'RITEX'),
    (r'C:\Users\MGaravaglia\Desktop\Control_Telas_Neocxela v2.xlsx', 'NEOCXELA'),
]

# Tipos de tela que no cuentan por piezas (solo kg)
TELAS_SIN_PIEZAS = {'MORLEY', 'RIBB', 'REEB'}


def buscar_proveedor(nombre_frag):
    """Busca un proveedor cuyo nombre contenga el fragmento (case-insensitive)."""
    p = Proveedor.query.filter(
        Proveedor.nombre.ilike(f'%{nombre_frag}%')
    ).first()
    return p


def importar_archivo(path, prov_frag):
    if not os.path.exists(path):
        print(f'[WARN] No existe: {path}')
        return 0, 0

    wb = load_workbook(path, data_only=True)
    if 'Maestro_Telas' not in wb.sheetnames:
        print(f'[WARN] {path} no tiene hoja Maestro_Telas')
        return 0, 0

    ws = wb['Maestro_Telas']
    prov = buscar_proveedor(prov_frag)
    if not prov:
        print(f'[WARN] No encuentro el proveedor "{prov_frag}" en la tabla proveedores. Se crea.')
        prov = Proveedor(nombre=prov_frag.title(), categoria='Textiles')
        db.session.add(prov)
        db.session.flush()

    print(f'\n=== {path} -> Proveedor "{prov.nombre}" (id={prov.id}) ===')

    creadas = 0
    existentes = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # columnas A/B/C: Articulo / Codigo / Color
        if len(row) < 3:
            continue
        tipo, cod_art, color = row[0], row[1], row[2]
        if not tipo or not color:
            continue
        tipo = str(tipo).strip().upper()
        color = str(color).strip().upper()
        cod_art = str(cod_art).strip() if cod_art else None

        # La columna B del Excel viene como "codigoTela-codigoColor" (ej. "5138-5469").
        # Separar en cod_art (tela) y cod_color (color).
        cod_color = None
        if cod_art and '-' in cod_art:
            left, right = cod_art.split('-', 1)
            cod_art = left.strip() or None
            cod_color = right.strip() or None

        # Verificar si ya existe
        q = MaestroTela.query.filter_by(
            proveedor_id=prov.id,
            tipo_tela=tipo,
            color=color,
        )
        if cod_art:
            q = q.filter_by(cod_art=cod_art)
        if cod_color:
            q = q.filter_by(cod_color=cod_color)
        existente = q.first()
        if existente:
            existentes += 1
            continue

        cuenta_piezas = tipo not in TELAS_SIN_PIEZAS

        tela = MaestroTela(
            proveedor_id=prov.id,
            tipo_tela=tipo,
            cod_art=cod_art,
            cod_color=cod_color,
            color=color,
            cuenta_piezas=cuenta_piezas,
            activo=True,
        )
        db.session.add(tela)
        creadas += 1

    db.session.commit()
    print(f'   Creadas: {creadas}   Ya existentes: {existentes}')
    return creadas, existentes


def main():
    with app.app_context():
        print('IMPORTADOR MAESTRO DE TELAS')
        print('=' * 60)
        total_c = 0
        total_e = 0
        for path, prov in FUENTES:
            c, e = importar_archivo(path, prov)
            total_c += c
            total_e += e
        print('\n' + '=' * 60)
        print(f'TOTAL: {total_c} creadas, {total_e} ya existentes.')
        print(f'Tabla maestro_telas ahora tiene {MaestroTela.query.count()} filas.')


if __name__ == '__main__':
    main()
