from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def excel_style(ws, headers, col_widths=None):
    """Aplica estilos al header de una hoja Excel."""
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='2C3E50')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    Border(bottom=thin)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    if col_widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
