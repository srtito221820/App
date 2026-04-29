import io
from datetime import date

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func

from models import (db, Proveedor, CuentaCorriente, MovimientoTela, Anticipo)
from blueprints._excel import excel_style

bp = Blueprint('exportar', __name__, url_prefix='/exportar')


@bp.route('/cuenta-corriente/<int:proveedor_id>')
def exportar_cc(proveedor_id):
    prov = Proveedor.query.get_or_404(proveedor_id)
    movs = CuentaCorriente.query.filter_by(proveedor_id=proveedor_id)\
        .order_by(CuentaCorriente.fecha, CuentaCorriente.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cuenta Corriente'
    headers = ['Fecha', 'Cuenta', 'Tipo', 'Comprobante', 'Descripcion', 'Debe', 'Haber', 'Saldo']
    excel_style(ws, headers, [14, 12, 16, 18, 40, 16, 16, 16])

    saldo = 0
    for i, m in enumerate(movs, 2):
        saldo += m.debe - m.haber
        ws.cell(row=i, column=1, value=m.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=2, value=m.cuenta or '')
        ws.cell(row=i, column=3, value=m.tipo or '')
        ws.cell(row=i, column=4, value=m.numero_comprobante or '')
        ws.cell(row=i, column=5, value=m.descripcion or '')
        ws.cell(row=i, column=6, value=m.debe).number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=m.haber).number_format = '#,##0.00'
        ws.cell(row=i, column=8, value=saldo).number_format = '#,##0.00'

    row = len(movs) + 2
    ws.cell(row=row, column=5, value='TOTALES').font = Font(bold=True)
    ws.cell(row=row, column=6, value=sum(m.debe for m in movs)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7, value=sum(m.haber for m in movs)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=8, value=saldo).number_format = '#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = prov.nombre.replace(' ', '_')
    return send_file(output, download_name=f'CuentaCorriente_{nombre}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/stock')
def exportar_stock():
    proveedor_id = request.args.get('proveedor_id', type=int)
    tipo_tela = request.args.get('tipo_tela', '').strip()
    color_f = request.args.get('color', '').strip()
    cod_art_f = request.args.get('cod_art', '').strip()
    cod_color_f = request.args.get('cod_color', '').strip()

    query = db.session.query(
        MovimientoTela.tipo_tela,
        MovimientoTela.color,
        MovimientoTela.cod_art,
        MovimientoTela.cod_color,
        Proveedor.nombre.label('proveedor'),
        func.sum(MovimientoTela.cant_kg).label('total_kg'),
        func.sum(MovimientoTela.piezas).label('total_piezas'),
    ).join(Proveedor).group_by(
        MovimientoTela.tipo_tela, MovimientoTela.color,
        MovimientoTela.cod_art, MovimientoTela.cod_color,
        Proveedor.nombre
    )
    if proveedor_id:
        query = query.filter(MovimientoTela.proveedor_id == proveedor_id)
    if tipo_tela:
        query = query.filter(MovimientoTela.tipo_tela == tipo_tela)
    if color_f:
        query = query.filter(MovimientoTela.color == color_f)
    if cod_art_f:
        query = query.filter(MovimientoTela.cod_art == cod_art_f)
    if cod_color_f:
        query = query.filter(MovimientoTela.cod_color == cod_color_f)
    stock = query.order_by(MovimientoTela.tipo_tela, MovimientoTela.color).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock de Tela'
    headers = ['Tipo Tela', 'Color', 'Cod Art', 'Cod Color', 'Proveedor', 'Stock (Kg)', 'Piezas']
    excel_style(ws, headers, [16, 18, 12, 12, 20, 14, 10])

    for i, s in enumerate(stock, 2):
        ws.cell(row=i, column=1, value=s.tipo_tela or '')
        ws.cell(row=i, column=2, value=s.color or '')
        ws.cell(row=i, column=3, value=s.cod_art or '')
        ws.cell(row=i, column=4, value=s.cod_color or '')
        ws.cell(row=i, column=5, value=s.proveedor or '')
        ws.cell(row=i, column=6, value=s.total_kg or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=s.total_piezas or 0)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name=f'Stock_Tela_{date.today().strftime("%Y%m%d")}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/movimientos-tela/<int:proveedor_id>')
def exportar_movimientos_tela(proveedor_id):
    prov = Proveedor.query.get_or_404(proveedor_id)
    q = MovimientoTela.query.filter_by(proveedor_id=proveedor_id)
    fecha_desde = request.args.get('tela_fecha_desde', '')
    fecha_hasta = request.args.get('tela_fecha_hasta', '')
    cuenta = request.args.get('tela_cuenta', '')
    remito = request.args.get('tela_remito', '')
    tipo_tela = request.args.get('tela_tipo_tela', '')
    color = request.args.get('tela_color', '')
    cod_art = request.args.get('tela_cod_art', '')
    temporada = request.args.get('tela_temporada', '')
    if fecha_desde:
        q = q.filter(MovimientoTela.fecha >= date.fromisoformat(fecha_desde))
    if fecha_hasta:
        q = q.filter(MovimientoTela.fecha <= date.fromisoformat(fecha_hasta))
    if cuenta:
        q = q.filter(MovimientoTela.cuenta == cuenta)
    if remito:
        q = q.filter(MovimientoTela.remito_factura == remito)
    if tipo_tela:
        q = q.filter(MovimientoTela.tipo_tela == tipo_tela)
    if color:
        q = q.filter(MovimientoTela.color == color)
    if cod_art:
        q = q.filter(MovimientoTela.cod_art == cod_art)
    if temporada:
        q = q.filter(MovimientoTela.temporada == temporada)
    movs = q.order_by(MovimientoTela.fecha, MovimientoTela.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos de Tela'
    headers = ['Fecha', 'Cuenta', 'Remito/Factura', 'Tipo Tela', 'Color',
               'Cod Art', 'Cod Color', 'Kg', 'Piezas', 'Partida',
               '$ s/IVA', '$ c/IVA', 'Subtotal', 'Percp IVA', 'Percp IIBB',
               'Sub c/IVA', 'Movimiento', 'Temporada', 'Anticipo', 'Pedido']
    excel_style(ws, headers, [14, 10, 16, 14, 16, 10, 10, 12, 8, 10, 12, 12, 14, 12, 12, 14, 12, 12, 14, 12])

    for i, m in enumerate(movs, 2):
        ws.cell(row=i, column=1, value=m.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=2, value=m.cuenta or '')
        ws.cell(row=i, column=3, value=m.remito_factura or '')
        ws.cell(row=i, column=4, value=m.tipo_tela or '')
        ws.cell(row=i, column=5, value=m.color or '')
        ws.cell(row=i, column=6, value=m.cod_art or '')
        ws.cell(row=i, column=7, value=m.cod_color or '')
        ws.cell(row=i, column=8, value=m.cant_kg or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=9, value=m.piezas or 0)
        ws.cell(row=i, column=10, value=m.partida or '')
        ws.cell(row=i, column=11, value=m.precio_sin_iva or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=12, value=m.precio_con_iva or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=13, value=m.subtotal or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=14, value=m.percp_iva or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=15, value=m.percp_iibb or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=16, value=m.subtotal_iva or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=17, value=m.movimiento or '')
        ws.cell(row=i, column=18, value=m.temporada or '')
        ws.cell(row=i, column=19, value=m.anticipo.numero if m.anticipo else '')
        ws.cell(row=i, column=20, value=m.pedido.numero if m.pedido else '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = prov.nombre.replace(' ', '_')
    return send_file(output, download_name=f'Movimientos_Tela_{nombre}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/anticipos/<int:proveedor_id>')
def exportar_anticipos(proveedor_id):
    prov = Proveedor.query.get_or_404(proveedor_id)
    anticipos = Anticipo.query.filter_by(proveedor_id=proveedor_id)\
        .order_by(Anticipo.fecha).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Anticipos'
    headers = ['Numero', 'Fecha', 'Factura', 'Estado', 'Monto', 'Kg Anticipo',
               'Kg Pedidos', 'Kg Entregados', 'Kg Pendientes', 'Saldo Pendiente']
    excel_style(ws, headers, [16, 14, 16, 12, 16, 14, 14, 14, 14, 16])

    for i, a in enumerate(anticipos, 2):
        ws.cell(row=i, column=1, value=a.numero)
        ws.cell(row=i, column=2, value=a.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=3, value=a.numero_factura or '')
        ws.cell(row=i, column=4, value=a.estado)
        ws.cell(row=i, column=5, value=a.monto or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=a.cant_kg or 0).number_format = '#,##0.00'
        ws.cell(row=i, column=7, value=a.total_kg_pedidos()).number_format = '#,##0.00'
        ws.cell(row=i, column=8, value=a.total_kg_entregados()).number_format = '#,##0.00'
        ws.cell(row=i, column=9, value=a.kg_pendientes()).number_format = '#,##0.00'
        ws.cell(row=i, column=10, value=a.saldo_pendiente()).number_format = '#,##0.00'

    ws2 = wb.create_sheet('Pedidos')
    headers2 = ['Anticipo', 'Pedido', 'Fecha', 'Estado', 'Tipo Tela', 'Color',
                'Cod Art', 'Kg Pedido', 'Kg Entregado', 'Kg Pendiente', '$/kg', 'Subtotal']
    excel_style(ws2, headers2, [16, 12, 14, 12, 14, 16, 10, 14, 14, 14, 12, 14])

    row = 2
    for a in anticipos:
        for p in a.pedidos.all():
            for d in p.detalles.all():
                ws2.cell(row=row, column=1, value=a.numero)
                ws2.cell(row=row, column=2, value=p.numero)
                ws2.cell(row=row, column=3, value=p.fecha.strftime('%d/%m/%Y'))
                ws2.cell(row=row, column=4, value=p.estado)
                ws2.cell(row=row, column=5, value=d.tipo_tela or '')
                ws2.cell(row=row, column=6, value=d.color or '')
                ws2.cell(row=row, column=7, value=d.cod_art or '')
                ws2.cell(row=row, column=8, value=d.cant_kg or 0).number_format = '#,##0.00'
                ws2.cell(row=row, column=9, value=d.kg_entregados()).number_format = '#,##0.00'
                ws2.cell(row=row, column=10, value=d.kg_pendientes()).number_format = '#,##0.00'
                ws2.cell(row=row, column=11, value=d.precio_unitario or 0).number_format = '#,##0.00'
                ws2.cell(row=row, column=12, value=d.subtotal or 0).number_format = '#,##0.00'
                row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = prov.nombre.replace(' ', '_')
    return send_file(output, download_name=f'Anticipos_{nombre}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
