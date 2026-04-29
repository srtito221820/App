import io
from datetime import date, datetime as _dt

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file)
from openpyxl import Workbook

from models import (db, Proveedor, MovimientoTela, MaestroTela, Partida)
from utils import registrar_auditoria
from blueprints._excel import excel_style

bp = Blueprint('consumos', __name__, url_prefix='/consumos')


def _parse_fecha(s):
    try:
        return _dt.strptime(s, '%Y-%m-%d').date() if s else None
    except ValueError:
        return None


def _consumos_filters_from_request():
    return {
        'proveedor_id': request.args.get('proveedor_id', type=int),
        'tipo_tela': request.args.get('tipo_tela', '').strip(),
        'color': request.args.get('color', '').strip(),
        'cuenta': request.args.get('cuenta', '').strip(),
        'partida_num': request.args.get('partida', '').strip(),
        'fecha_desde': _parse_fecha(request.args.get('fecha_desde', '').strip()),
        'fecha_hasta': _parse_fecha(request.args.get('fecha_hasta', '').strip()),
        'incluir_agotadas': request.args.get('agotadas') == '1',
    }


def _aplicar_filtros_consumos_partidas(f):
    q = MovimientoTela.query.filter_by(movimiento='Consumo')
    if f['proveedor_id']:
        q = q.filter(MovimientoTela.proveedor_id == f['proveedor_id'])
    if f['tipo_tela']:
        q = q.filter(MovimientoTela.tipo_tela == f['tipo_tela'])
    if f['color']:
        q = q.filter(MovimientoTela.color == f['color'])
    if f['cuenta']:
        q = q.filter(MovimientoTela.cuenta == f['cuenta'])
    if f['partida_num']:
        q = q.join(Partida, MovimientoTela.partida_id == Partida.id) \
             .filter(Partida.numero == f['partida_num'])
    if f['fecha_desde']:
        q = q.filter(MovimientoTela.fecha >= f['fecha_desde'])
    if f['fecha_hasta']:
        q = q.filter(MovimientoTela.fecha <= f['fecha_hasta'])
    consumos = q.order_by(MovimientoTela.fecha.desc(), MovimientoTela.id.desc()).all()

    pq = Partida.query.filter_by(activo=True)
    if f['proveedor_id']:
        pq = pq.filter(Partida.proveedor_id == f['proveedor_id'])
    if f['tipo_tela']:
        pq = pq.filter(Partida.tipo_tela == f['tipo_tela'])
    if f['color']:
        pq = pq.filter(Partida.color == f['color'])
    if f['partida_num']:
        pq = pq.filter(Partida.numero == f['partida_num'])
    if f['fecha_desde']:
        pq = pq.filter(Partida.fecha_alta >= f['fecha_desde'])
    if f['fecha_hasta']:
        pq = pq.filter(Partida.fecha_alta <= f['fecha_hasta'])
    partidas = pq.order_by(Partida.fecha_alta.desc(), Partida.numero.desc()).all()
    if not f['incluir_agotadas']:
        partidas = [p for p in partidas if not p.esta_agotada()]
    return consumos, partidas


@bp.route('')
def consumos_list():
    f = _consumos_filters_from_request()
    consumos, partidas = _aplicar_filtros_consumos_partidas(f)
    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.nombre).all()

    total_kg = sum((m.cant_kg or 0) for m in consumos)
    total_piezas = sum((m.piezas or 0) for m in consumos)

    def _distinct(sources):
        s = set()
        for src in sources:
            for (v,) in src:
                if v:
                    s.add(v)
        return sorted(s)

    tipos_tela = _distinct([
        db.session.query(MovimientoTela.tipo_tela).filter_by(movimiento='Consumo').distinct(),
        db.session.query(Partida.tipo_tela).filter_by(activo=True).distinct(),
    ])
    colores = _distinct([
        db.session.query(MovimientoTela.color).filter_by(movimiento='Consumo').distinct(),
        db.session.query(Partida.color).filter_by(activo=True).distinct(),
    ])
    partidas_nums = _distinct([
        db.session.query(Partida.numero).filter_by(activo=True).distinct(),
    ])

    return render_template('consumos/list.html',
                           consumos=consumos, proveedores=proveedores,
                           tipos_tela=tipos_tela, colores=colores,
                           partidas_nums=partidas_nums, partidas=partidas,
                           proveedor_id=f['proveedor_id'], tipo_tela_sel=f['tipo_tela'],
                           color_sel=f['color'], cuenta_sel=f['cuenta'],
                           partida_sel=f['partida_num'],
                           fecha_desde_sel=request.args.get('fecha_desde', ''),
                           fecha_hasta_sel=request.args.get('fecha_hasta', ''),
                           incluir_agotadas=f['incluir_agotadas'],
                           total_kg=total_kg, total_piezas=total_piezas)


@bp.route('/exportar')
def exportar_consumos():
    f = _consumos_filters_from_request()
    consumos, partidas = _aplicar_filtros_consumos_partidas(f)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Partidas'
    headers1 = ['Nº Partida', 'Fecha Alta', 'Proveedor', 'Tipo Tela', 'Color',
                'Pzas Totales', 'Pzas Consumidas', 'Pzas Saldo',
                'Kg Totales', 'Kg Consumidos', 'Kg Saldo', 'Estado']
    excel_style(ws1, headers1, [12, 12, 22, 14, 14, 12, 14, 12, 12, 14, 12, 12])
    for i, p in enumerate(partidas, 2):
        ws1.cell(row=i, column=1, value=p.numero or '')
        ws1.cell(row=i, column=2, value=p.fecha_alta).number_format = 'DD/MM/YYYY'
        ws1.cell(row=i, column=3, value=p.proveedor.nombre if p.proveedor else '')
        ws1.cell(row=i, column=4, value=p.tipo_tela or '')
        ws1.cell(row=i, column=5, value=p.color or '')
        ws1.cell(row=i, column=6, value=p.piezas_totales or 0)
        ws1.cell(row=i, column=7, value=p.piezas_consumidas())
        ws1.cell(row=i, column=8, value=p.piezas_saldo())
        ws1.cell(row=i, column=9, value=p.kg_totales or 0).number_format = '#,##0.00'
        ws1.cell(row=i, column=10, value=p.kg_consumidos()).number_format = '#,##0.00'
        ws1.cell(row=i, column=11, value=p.kg_saldo()).number_format = '#,##0.00'
        ws1.cell(row=i, column=12, value='Agotada' if p.esta_agotada() else 'Activa')

    ws2 = wb.create_sheet('Consumos')
    headers2 = ['Fecha', 'Proveedor', 'Tipo Tela', 'Color', 'Partida',
                'OP', 'Kg', 'Piezas', 'Cuenta', 'Observaciones']
    excel_style(ws2, headers2, [12, 22, 14, 14, 12, 14, 10, 10, 12, 32])
    for i, m in enumerate(consumos, 2):
        ws2.cell(row=i, column=1, value=m.fecha).number_format = 'DD/MM/YYYY'
        ws2.cell(row=i, column=2, value=m.proveedor.nombre if m.proveedor else '')
        ws2.cell(row=i, column=3, value=m.tipo_tela or '')
        ws2.cell(row=i, column=4, value=m.color or '')
        ws2.cell(row=i, column=5, value=(m.partida_rel.numero if m.partida_rel else (m.partida or '')))
        ws2.cell(row=i, column=6, value=m.op or '')
        ws2.cell(row=i, column=7, value=m.cant_kg or 0).number_format = '#,##0.00'
        ws2.cell(row=i, column=8, value=m.piezas or 0)
        ws2.cell(row=i, column=9, value=m.cuenta or '')
        ws2.cell(row=i, column=10, value=m.observaciones or '')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name=f'Consumos_{date.today().strftime("%Y%m%d")}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/nuevo', methods=['GET', 'POST'])
def consumo_nuevo():
    if request.method == 'POST':
        try:
            proveedor_id = request.form.get('proveedor_id', type=int)
            maestro_tela_id = request.form.get('maestro_tela_id', type=int) or None
            partida_id = request.form.get('partida_id', type=int) or None
            cant_kg = request.form.get('cant_kg', type=float) or 0
            piezas = request.form.get('piezas', type=int) or 0
            cuenta = request.form.get('cuenta', '').strip() or None
            op = request.form.get('op', '').strip() or None
            fecha_str = request.form.get('fecha') or date.today().isoformat()
            obs = request.form.get('observaciones', '').strip() or None

            if not proveedor_id or cant_kg <= 0:
                flash('Proveedor y cantidad (kg) > 0 son obligatorios.', 'danger')
                raise ValueError('Datos invalidos')

            tela = db.session.get(MaestroTela, maestro_tela_id) if maestro_tela_id else None
            part = db.session.get(Partida, partida_id) if partida_id else None

            if part:
                if (part.piezas_totales or 0) > 0 and piezas > part.piezas_saldo():
                    flash(f'No se puede consumir {piezas} piezas: la partida {part.numero} '
                          f'tiene saldo de {part.piezas_saldo()} piezas.', 'danger')
                    raise ValueError('Piezas exceden saldo')
                if (part.kg_totales or 0) > 0 and cant_kg > part.kg_saldo() + 0.001:
                    flash(f'No se puede consumir {cant_kg:.2f} kg: la partida {part.numero} '
                          f'tiene saldo de {part.kg_saldo():.2f} kg.', 'danger')
                    raise ValueError('Kg exceden saldo')

            m = MovimientoTela(
                fecha=_dt.strptime(fecha_str, '%Y-%m-%d').date(),
                proveedor_id=proveedor_id,
                cuenta=cuenta,
                tipo_tela=(tela.tipo_tela if tela else (part.tipo_tela if part else '')),
                color=(tela.color if tela else (part.color if part else '')),
                cod_art=(tela.cod_art if tela else (part.cod_art if part else None)),
                cod_color=(tela.cod_color if tela else (part.cod_color if part else None)),
                descripcion=(tela.descripcion if tela else None),
                cant_kg=-abs(cant_kg),
                piezas=-abs(piezas),
                movimiento='Consumo',
                estado='Consumido',
                op=op,
                observaciones=obs,
                maestro_tela_id=maestro_tela_id,
                partida_id=partida_id,
            )
            db.session.add(m)
            db.session.flush()
            registrar_auditoria('CREAR', 'Consumo', m.id,
                                f'{m.tipo_tela}/{m.color} {cant_kg}kg')
            db.session.commit()
            flash('Consumo registrado.', 'success')
            return redirect(url_for('consumos.consumos_list'))
        except ValueError:
            pass
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')

    proveedores = Proveedor.query.filter_by(activo=True).order_by(Proveedor.nombre).all()
    partida_pre = None
    partida_id_pre = (request.args.get('partida_id', type=int)
                      or request.form.get('partida_id', type=int))
    if partida_id_pre:
        partida_pre = Partida.query.filter_by(id=partida_id_pre, activo=True).first()
    return render_template('consumos/form.html', consumo=None,
                           proveedores=proveedores, partida_pre=partida_pre)


@bp.route('/<int:id>/eliminar', methods=['POST'])
def consumo_eliminar(id):
    m = MovimientoTela.query.get_or_404(id)
    if m.movimiento != 'Consumo':
        flash('No es un consumo.', 'danger')
        return redirect(url_for('consumos.consumos_list'))
    registrar_auditoria('ELIMINAR', 'Consumo', m.id,
                        f'{m.tipo_tela}/{m.color} {m.cant_kg}kg')
    db.session.delete(m)
    db.session.commit()
    flash('Consumo eliminado.', 'info')
    return redirect(url_for('consumos.consumos_list'))
